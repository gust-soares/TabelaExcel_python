import openpyxl

book = openpyxl.Workbook()

print(book.sheetnames)

book.create_sheet('Frutas')

frutas_page = book['Frutas']
frutas_page.append(['Fruta', 'Quantidade', 'Preço'])

frutas_page.append(['Banana', '5', 'R$ 3,90'])
frutas_page.append(['Fruta 2', '5', 'R$ 15,90'])
frutas_page.append(['Fruta 3', '5', 'R$ 30,90'])
frutas_page.append(['Fruta 4', '2', 'R$ 50,50'])

book.save('Planilha de Compras.xlsx')
